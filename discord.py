import youtube_dl
import discord
import openpyxl

 
client = discord.Client()


@client.event
async def on_ready():
    print(client.user.id)
    print("ready")
    game = discord.Game("재깨봇 테스트중")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async def on_message(message):
    if message.content.startswith(",배워"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 101):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("단어가 학습되었습니다.")
                break
        file.save("기억.xlsx")

    if message.content.startswith("재깨봇"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 101):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break
    if message.content.startswith(",입장"):
        await message.author.voice.channel.connect()
        await message.channel.send("보이스 채널에 입장하였습니다.")
    if message.content.startswith(",퇴장"):
        for vc in client.voice_clients:
            if vc.guild == message.guild:
                voice = vc

        await voice.disconnect()
        await message.channel.send("보이스 채널을 퇴장하였습니다.")
    if message.content.startswith(",재생"):
        for vc in client.voice_clients:
            if vc.guild == message.guild:
                voice = vc
        url = message.content.split(" ")[1]
        option = {'format': 'bestaudio/best', 'postprocessors': [{'preferredcodec': 'mp3', 'key': 'FFmpegExtractAudio', 'preferredquality': '320',}], 'outtmpl' : "file/" + url.split('=')[1] + '.mp3'}

        with youtube_dl.YoutubeDL(option) as ydl:
            ydl.download([url])
            info = ydl.extract_info(url, download=False)
            title = info["title"]

        voice.play(discord.FFmpegPCMAudio("file/" + url.split('=')[1] + ".mp3"))
        await message.channel.send(title + "을 재생합니다.")

    if message.content.startswith("안녕"):
        await message.channel.send("안뇽!")
    if message.content.startswith(',명령어'):
        embed = discord.Embed(title="재깨봇 명령어",description="http://commands.jkbot.kro.kr", color=0x00aeee)
        await message.channel.send(embed=embed)
    if message.content.startswith(',오일카'):
        embed = discord.Embed(title="오버워치 일상카페",description="https://cafe.naver.com/redciqyi", color=0x00aeee)
        embed.add_field(name="현재 오일카에서 하는 이벤트 확인!", value=",이벤트", inline=False)
        await message.channel.send(embed=embed)
    if message.content.startswith(',초대'):
        embed = discord.Embed(title="재깨봇 초대",description="http://bot.jkbot.kro.kr", color=0x00aeee)
    if message.content.startswith(',재깨봇'):
        embed = discord.Embed(title="재깨봇 프로필",description="", color=0x00aeee)
        embed.add_field(name="재깨봇 생일", value="5월 28일!", inline=False)
        embed.add_field(name="재깨봇 제작자", value="유민#3285", inline=False)
        embed.add_field(name="재깨봇 사이트", value="https://jkbot.kro.kr", inline=False)
        await message.channel.send(embed=embed)
    if message.content.startswith(',이벤트'):
        embed = discord.Embed(title="오버워치 일상카페 이벤트",description="", color=0x00aeee)
        embed.add_field(name="오일카 내전 시즌 2!", value="https://cafe.naver.com/redciqyi/7592", inline=False)
        await message.channel.send(embed=embed)
    if message.content.startswith(",청소"):
        i = (message.author.guild_permissions.administrator)

        if i is True:
            amount = message.content[2:99]
            await message.channel.purge(limit=1)
            await message.channel.purge(limit=int(amount))

            embed = discord.Embed(title="메시지 삭제", description="디스코드 채팅 {}개를\n{}님이 삭제하였습니다.".format(amount, message.author), color=0x00aeee)
            await message.channel.send(embed=embed)
        
        if i is False:
            await message.channel.purge(limit=1)
            await message.channel.send("{}, 관리자 권한이 없어요!".format(message.author.mention))
    if message.content.startswith("") and message.author.id != 847399598419214368:
        file = openpyxl.load_workbook("레벨.xlsx")
        sheet = file.active
        exp = [10, 20, 30, 40, 50]
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                sheet["B" + str(i)].value = sheet["B" + str(i)].value + 1
                if sheet["B" + str(i)].value >= exp[sheet["C" + str(i)].value]:
                    sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                    await message.channel.send("레벨이 올랐어요! ₩n 현재 레벨 : " + str(sheet["C" + str(i)].value))
                file.save("레벨.xlsx")
                break
                 
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["A" + str(i)].value = 0
                sheet["A" + str(i)].value = 1
                file.save("레벨.xlsx")
                break

            i += 1



client.run("DISCORD_BOT_TOKEN")
